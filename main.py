import streamlit as st
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as ptl
from plotly.graph_objects import Figure, Scatter
import numpy as np
import seaborn as sns



st.title('Практическая рабора по анализу изменения зарплат на рынке труда')


#pd.read_excel('/content/tab3-zpl_2023.xlsx')
book = openpyxl.open("tab3-zpl_2023.xlsx", read_only=True)
sheet = book.active

# Create a list of column names
ccolumns = [sheet.cell(1, i).value for i in range(1, 9)]

# Create a list of data rows
data = [
    [sheet.cell(i, j).value for j in range(1, 9)]
    for i in range(45, 50)
]

# Create the DataFrame
table_for_analysys = pd.DataFrame(data, index=[sheet.cell(i, 1).value for i in range(45, 50)], columns=ccolumns)

st.table(table_for_analysys)





#Analysys of information and commication field
it_and_communication2017 = table_for_analysys.at['деятельность в области информации и связи', 2017.0]
it_and_communication2018 = table_for_analysys.at['деятельность в области информации и связи', 2018.0]
it_and_communication2019 = table_for_analysys.at['деятельность в области информации и связи', 2019.0]
it_and_communication2020 = table_for_analysys.at['деятельность в области информации и связи', 2020.0]
it_and_communication2021 = table_for_analysys.at['деятельность в области информации и связи', 2021.0]
it_and_communication2022 = table_for_analysys.at['деятельность в области информации и связи', 2022.0]
it_and_communication2023 = table_for_analysys.at['деятельность в области информации и связи', 2023.0]

list_of_job_salary = pd.Series({"2017": it_and_communication2017, "2018":  it_and_communication2018, "2019": it_and_communication2019, "2020": it_and_communication2020, "2021": it_and_communication2021, "2022": it_and_communication2022, "2023": it_and_communication2023})
#ptl.plot(list_of_job_salary)
#ptl.show()
#print('Изменение зарплат в сфере информации и связи:')
fig1_it = Figure(data=Scatter(x=list_of_job_salary.index, y=list_of_job_salary.values))
#fig.show()
fig1_it.update_layout(title_text='Изменение зарплат в сфере информации и связи:')

st.write(fig1_it)



#Analysys of science and technology
science_and_technology2017 = table_for_analysys.at['деятельность профессиональная,научная и техническая', 2017.0]
science_and_technology2018 = table_for_analysys.at['деятельность профессиональная,научная и техническая', 2018.0]
science_and_technology2019 = table_for_analysys.at['деятельность профессиональная,научная и техническая', 2019.0]
science_and_technology2020 = table_for_analysys.at['деятельность профессиональная,научная и техническая', 2020.0]
science_and_technology2021 = table_for_analysys.at['деятельность профессиональная,научная и техническая', 2021.0]
science_and_technology2022 = table_for_analysys.at['деятельность профессиональная,научная и техническая', 2022.0]
science_and_technology2023 = table_for_analysys.at['деятельность профессиональная,научная и техническая', 2023.0]

list_of_job_salary_science_tech = pd.Series({"2017": science_and_technology2017, "2018":  science_and_technology2018, "2019": science_and_technology2019, "2020": science_and_technology2020, "2021": science_and_technology2021, "2022": science_and_technology2022, "2023": science_and_technology2023})

#print('Изменение зарплат в сфере науки и техногий:')
#ptl.titleline('Изменение зарплат в сфере науки и техногий:')
fig_sci = Figure(data=Scatter(x=list_of_job_salary_science_tech.index, y=list_of_job_salary_science_tech.values))
#fig.show()
fig_sci.update_layout(title_text='Изменение зарплат в сфере науки и техногий:')
#fig_sci.show()

st.write(fig_sci)



#Analysys of science and technology and deveping inventions
science_and_technology_dev2017 = table_for_analysys.at['  из нее научные исследования и разработки', 2017.0]
science_and_technology_dev2018 = table_for_analysys.at['  из нее научные исследования и разработки', 2018.0]
science_and_technology_dev2019 = table_for_analysys.at['  из нее научные исследования и разработки', 2019.0]
science_and_technology_dev2020 = table_for_analysys.at['  из нее научные исследования и разработки', 2020.0]
science_and_technology_dev2021 = table_for_analysys.at['  из нее научные исследования и разработки', 2021.0]
science_and_technology_dev2022 = table_for_analysys.at['  из нее научные исследования и разработки', 2022.0]
science_and_technology_dev2023 = table_for_analysys.at['  из нее научные исследования и разработки', 2023.0]

list_of_job_salary_science_tech_and_dev = pd.Series({"2017": science_and_technology_dev2017, "2018":  science_and_technology_dev2018, "2019": science_and_technology_dev2019, "2020": science_and_technology_dev2020, "2021": science_and_technology_dev2021, "2022": science_and_technology_dev2022, "2023": science_and_technology_dev2023})

#print('Изменение зарплат в сфере разарботки и инноваций:')
#ptl.titleline('Изменение зарплат в сфере разарботки и инноваций:')
fig_sci_dev = Figure(data=Scatter(x=list_of_job_salary_science_tech_and_dev.index, y=list_of_job_salary_science_tech_and_dev.values))

fig_sci_dev .update_layout(title_text='Изменение зарплат в сфере разарботки и инноваций:')

st.write(fig_sci_dev )



#Creating visual graph of inflation
inflation = pd.Series({"2017": 2.52, "2018": 4.27, "2019": 3.05, "2020": 4.91, "2021": 8.39, "2022": 11.92, "2023": 7.42, "2024": 1.55})

#print('Изменение Инфляции в процентах с 2017 по 2023 года:')
#ptl.titleline('Изменение Инфляции в процентах с 2017 по 2023 года:')
fig_inflat = Figure(data=Scatter(x=inflation.index, y=inflation.values))
#fig.show()
fig_inflat.update_layout(title_text='Изменение Инфляции в процентах с 2017 по 2023 года:')
#fig_inflat.show()

st.write(fig_inflat )




#Calculate salary with inflation
#Salary in it and communication with inflation
#Formula which is used salary_with_inflation = salary2017 * (inflation2018 + 1) * (inflation2019 + 1) * (inflation2020 + 1)...
#Another better formula was found: real salary = salary - (salary/100 * inflation)

#variables describes inflation
inflation2017 = inflation["2017"]
inflation2018 = inflation["2018"]
inflation2019 = inflation["2019"]
inflation2020 = inflation["2020"]
inflation2021 = inflation["2021"]
inflation2022 = inflation["2022"]
inflation2023 = inflation["2023"]
inflation2024 = inflation["2024"]


#salary in information and communication
salary_it2017 = it_and_communication2017 - (it_and_communication2017/100 * inflation2017)
salary_it2018 = it_and_communication2018 - (it_and_communication2018/100 * inflation2018)
salary_it2019 = it_and_communication2019 - (it_and_communication2019/100 * inflation2019)
salary_it2020 = it_and_communication2020 - (it_and_communication2020/100 * inflation2020)
salary_it2021 = it_and_communication2021 - (it_and_communication2021/100 * inflation2021)
salary_it2022 = it_and_communication2022 - (it_and_communication2022/100 * inflation2022)
salary_it2023 = it_and_communication2023 - (it_and_communication2023/100 * inflation2023)

list_of_job_salary = pd.Series({"2017": salary_it2017, "2018":  salary_it2018, "2019": salary_it2019, "2020": salary_it2020, "2021": salary_it2021, "2022": salary_it2022, "2023": salary_it2023})

fig1_it_with_iflation = Figure(data=Scatter(x=list_of_job_salary.index, y=list_of_job_salary.values))

fig1_it_with_iflation.update_layout(title_text='Изменение зарплат в сфере информации и связи с учетом инфляции:')
#fig1_it_with_iflation.show()

st.write(fig1_it_with_iflation)




#Salary in science and technology
salary_sci_tech2017 = science_and_technology2017 - (science_and_technology2017/100 * inflation2017)
salary_sci_tech2018 = science_and_technology2018 - (science_and_technology2018/100 * inflation2018)
salary_sci_tech2019 = science_and_technology2019 - (science_and_technology2019/100 * inflation2019)
salary_sci_tech2020 = science_and_technology2020 - (science_and_technology2020/100 * inflation2020)
salary_sci_tech2021 = science_and_technology2021 - (science_and_technology2021/100 * inflation2021)
salary_sci_tech2022 = science_and_technology2022 - (science_and_technology2022/100 * inflation2022)
salary_sci_tech2023 = science_and_technology2023 - (science_and_technology2023/100 * inflation2023)

list_of_job_salary = pd.Series({"2017": salary_sci_tech2017, "2018":  salary_sci_tech2018, "2019": salary_sci_tech2019, "2020": salary_sci_tech2020, "2021": salary_sci_tech2021, "2022": salary_sci_tech2022, "2023": salary_sci_tech2023})

fig1_sci_tech_with_inflation = Figure(data=Scatter(x=list_of_job_salary.index, y=list_of_job_salary.values))

fig1_sci_tech_with_inflation.update_layout(title_text='Изменение зарплат в сфере науки и технологий с учетом инфляции:')
#fig1_sci_tech_with_inflation.show()
st.write(fig1_sci_tech_with_inflation)


salary_science_dev2017 = science_and_technology_dev2017 - (science_and_technology_dev2017/100 * inflation2017)
salary_science_dev2018 = science_and_technology_dev2018 - (science_and_technology_dev2018/100 * inflation2018)
salary_science_dev2019 = science_and_technology_dev2019 - (science_and_technology_dev2019/100 * inflation2019)
salary_science_dev2020 = science_and_technology_dev2020 - (science_and_technology_dev2020/100 * inflation2020)
salary_science_dev2021 = science_and_technology_dev2021 - (science_and_technology_dev2021/100 * inflation2021)
salary_science_dev2022 = science_and_technology_dev2022 - (science_and_technology_dev2022/100 * inflation2022)
salary_science_dev2023 = science_and_technology_dev2023 - (science_and_technology_dev2023/100 * inflation2023)

list_of_job_salary = pd.Series({"2017": salary_science_dev2017, "2018":  salary_science_dev2018, "2019": salary_science_dev2019, "2020": salary_science_dev2020, "2021": salary_science_dev2021, "2022": salary_science_dev2022, "2023": salary_science_dev2023})

fig1_sci_tech_dev_with_inflation = Figure(data=Scatter(x=list_of_job_salary.index, y=list_of_job_salary.values))

fig1_sci_tech_dev_with_inflation.update_layout(title_text='Изменение зарплат в сфере науки, разработки и инноваций с учетом инфляции:')
#ffig1_sci_tech_dev_with_inflation.show()
st.write(fig1_sci_tech_dev_with_inflation)



# Calculation real salary index

st.subheader('Индекс цен и индексы номинальной и реальной заработной платы')


#index of consumption costs

index_of_consumption_costs = pd.Series({"2017": 104.26, "2018": 103.04, "2019": 104.91 , "2020": 108.39 , "2021": 111.94 , "2022": 107.42 , "2023": 101.95})
st.text('Таблица индексов цен')
st.table(index_of_consumption_costs)


#salary in information and communication
#inom_salary_it2017 = it_and_communication2017 - (it_and_communication2017/100 * inflation2017)
salary_it_ind2018 = it_and_communication2018 / (it_and_communication2017/100)
salary_it_ind2019 = it_and_communication2019 / (it_and_communication2018/100)
salary_it_ind2020 = it_and_communication2020 / (it_and_communication2019/100)
salary_it_ind2021 = it_and_communication2021 / (it_and_communication2020/100)
salary_it_ind2022 = it_and_communication2022 / (it_and_communication2021/100)
salary_it_ind2023 = it_and_communication2023 / (it_and_communication2022/100)

list_of_job_salary = pd.Series({"2018":  salary_it_ind2018, "2019": salary_it_ind2019, "2020": salary_it_ind2020, "2021": salary_it_ind2021, "2022": salary_it_ind2022, "2023": salary_it_ind2023})

fig1_it_with_inom = Figure(data=Scatter(x=list_of_job_salary.index, y=list_of_job_salary.values))

fig1_it_with_inom.update_layout(title_text='Индекс Номинальной заработной платы в сфере информации и связи')

ind_real_slr_it2018 =  salary_it_ind2018 / index_of_consumption_costs["2018"]
ind_real_slr_it2019 =  salary_it_ind2019 / index_of_consumption_costs["2019"]
ind_real_slr_it2020 =  salary_it_ind2020 / index_of_consumption_costs["2020"]
ind_real_slr_it2021 =  salary_it_ind2021 / index_of_consumption_costs["2021"]
ind_real_slr_it2022 =  salary_it_ind2022 / index_of_consumption_costs["2022"]
ind_real_slr_it2023 =  salary_it_ind2023 / index_of_consumption_costs["2023"]

list_of_job_salary_ind = pd.Series({"2018":  ind_real_slr_it2018, "2019": ind_real_slr_it2019, "2020": ind_real_slr_it2020, "2021": ind_real_slr_it2021, "2022": ind_real_slr_it2022, "2023": ind_real_slr_it2023})
list_of_job_salary_index = Figure(data=Scatter(x=list_of_job_salary_ind.index, y=list_of_job_salary_ind.values))

list_of_job_salary_index.update_layout(title_text='Индекс Реальной заработной платы в сфере информации и связи')

st.write(fig1_it_with_inom)
st.write(list_of_job_salary_index)






#index of salary in science and technology


salary_sci_tech_ind2018 = science_and_technology2018 / (science_and_technology2017/100)
salary_sci_tech_ind2019 = science_and_technology2019 / (science_and_technology2018/100)
salary_sci_tech_ind2020 = science_and_technology2020 / (science_and_technology2019/100)
salary_sci_tech_ind2021 = science_and_technology2021 / (science_and_technology2020/100)
salary_sci_tech_ind2022 = science_and_technology2022 / (science_and_technology2021/100)
salary_sci_tech_ind2023 = science_and_technology2023 / (science_and_technology2022/100)

list_of_job_salary_ind_sci_tech = pd.Series({"2018":  salary_sci_tech_ind2018, "2019": salary_sci_tech_ind2019, "2020": salary_sci_tech_ind2020, "2021": salary_sci_tech_ind2021, "2022": salary_sci_tech_ind2022, "2023": salary_sci_tech_ind2023})

fig1_sci_with_inom = Figure(data=Scatter(x=list_of_job_salary.index, y=list_of_job_salary.values))

fig1_sci_with_inom.update_layout(title_text='Индекс Номинальной заработной платы в сфере науки и технологий')

ind_real_slr_sci_tech2018 =  salary_sci_tech_ind2018 / index_of_consumption_costs["2018"]
ind_real_slr_sci_tech2019 =  salary_sci_tech_ind2019 / index_of_consumption_costs["2019"]
ind_real_slr_sci_tech2020 =  salary_sci_tech_ind2020 / index_of_consumption_costs["2020"]
ind_real_slr_sci_tech2021 =  salary_sci_tech_ind2021 / index_of_consumption_costs["2021"]
ind_real_slr_sci_tech2022 =  salary_sci_tech_ind2022 / index_of_consumption_costs["2022"]
ind_real_slr_sci_tech2023 =  salary_sci_tech_ind2023 / index_of_consumption_costs["2023"]

list_of_job_salary_sci_real_ind = pd.Series({"2018":  ind_real_slr_sci_tech2018, "2019": ind_real_slr_sci_tech2019, "2020": ind_real_slr_sci_tech2020, "2021": ind_real_slr_sci_tech2021, "2022": ind_real_slr_sci_tech2022, "2023": ind_real_slr_sci_tech2023})
list_of_job_salary_graph_index = Figure(data=Scatter(x=list_of_job_salary_ind.index, y=list_of_job_salary_ind.values))

list_of_job_salary_graph_index.update_layout(title_text='Индекс Реальной заработной платы в сфере науки и технологий')

st.write(fig1_sci_with_inom)
st.write(list_of_job_salary_graph_index)





# Index of salary in science developing and innovations

salary_sci_tech_dev_ind2018 = science_and_technology_dev2018 / (science_and_technology_dev2017/100)
salary_sci_tech_dev_ind2019 = science_and_technology_dev2019 / (science_and_technology_dev2018/100)
salary_sci_tech_dev_ind2020 = science_and_technology_dev2020 / (science_and_technology_dev2019/100)
salary_sci_tech_dev_ind2021 = science_and_technology_dev2021 / (science_and_technology_dev2020/100)
salary_sci_tech_dev_ind2022 = science_and_technology_dev2022 / (science_and_technology_dev2021/100)
salary_sci_tech_dev_ind2023 = science_and_technology_dev2023 / (science_and_technology_dev2022/100)

list_of_job_salary_ind_sci_tech_dev = pd.Series({"2018":  salary_sci_tech_dev_ind2018, "2019": salary_sci_tech_dev_ind2019, "2020": salary_sci_tech_dev_ind2020, "2021": salary_sci_tech_dev_ind2021, "2022": salary_sci_tech_dev_ind2022, "2023": salary_sci_tech_dev_ind2023})

fig1_sci_dev_with_inom = Figure(data=Scatter(x=list_of_job_salary.index, y=list_of_job_salary.values))

fig1_sci_dev_with_inom.update_layout(title_text='Индекс Номинальной заработной платы в сфере науки, разработки и инноваций')

ind_real_slr_sci_tech_dev2018 =  salary_sci_tech_dev_ind2018 / index_of_consumption_costs["2018"]
ind_real_slr_sci_tech_dev2019 =  salary_sci_tech_dev_ind2019 / index_of_consumption_costs["2019"]
ind_real_slr_sci_tech_dev2020 =  salary_sci_tech_dev_ind2020 / index_of_consumption_costs["2020"]
ind_real_slr_sci_tech_dev2021 =  salary_sci_tech_dev_ind2021 / index_of_consumption_costs["2021"]
ind_real_slr_sci_tech_dev2022 =  salary_sci_tech_dev_ind2022 / index_of_consumption_costs["2022"]
ind_real_slr_sci_tech_dev2023 =  salary_sci_tech_dev_ind2023 / index_of_consumption_costs["2023"]

list_of_job_salary_sci_dev_real_ind = pd.Series({"2018":  ind_real_slr_sci_tech_dev2018, "2019": ind_real_slr_sci_tech_dev2019, "2020": ind_real_slr_sci_tech_dev2020, "2021": ind_real_slr_sci_tech_dev2021, "2022": ind_real_slr_sci_tech_dev2022, "2023": ind_real_slr_sci_tech_dev2023})
list_of_job_salary_sci_dev_graph_index = Figure(data=Scatter(x=list_of_job_salary_ind.index, y=list_of_job_salary_ind.values))

list_of_job_salary_sci_dev_graph_index.update_layout(title_text='Индекс Реальной заработной платы в сфере науки, разработки и инноваций')

st.write(fig1_sci_dev_with_inom)
st.write(list_of_job_salary_sci_dev_graph_index)
























